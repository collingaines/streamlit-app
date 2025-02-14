#=====================================================================================================================================================================================================================================================================================================================
#=====================================================================================================================================================================================================================================================================================================================
#SECTION #0 | IMPORTING OUR PYTHON LIBRARIES: 
#region CLICK HERE TO EXPAND SECTION


#===============================================================================================================================================================
#Importing misc libraries
#region

print('IMPORTING PYTHON LIBRARIES...')

import pandas as pd
import numpy as np
import datetime

#The time module will alow us to pause the script for set periods of time. This can be useful when some information takes a couple of seconds to update properly on the webpage
import time

#importing the os module to allow us to work with our operating system in various ways
import os

#Import sqlite3 for all database functionality
import sqlite3

#importing openpyxl for excel interaction, including modules that allow us to incorporate conditional formatting:
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles import Border, Side

import io

print('SUCCESS')

#endregion

#===============================================================================================================================================================
#Defining some misc functions for use later in script:
#region

print('DEFINING MISC FUNCTIONS FOR USE THROUGHOUT SCRIPT...')

#================================================================
#Defining our function to pull an address from GPS coordinates:
from geopy.geocoders import Nominatim
from geopy.exc import GeocoderTimedOut

def get_address_from_coordinates(latitude, longitude):
    """
    Takes GPS coordinates (latitude, longitude) and returns the corresponding address.
    """
    geolocator = Nominatim(user_agent="geo_lookup")
            
    try:
        location = geolocator.reverse((latitude, longitude), exactly_one=True)
        return location.address if location else "Address not found"
    except GeocoderTimedOut:
        return "Geocoder service timed out. Try again."

#================================================================
#Writing a function that will format our code block run time print outs in the console to be easier to read: 
def format_time(seconds):
    minutes = int(seconds // 60)
    seconds = seconds % 60
    return f"{minutes} minutes and {seconds:.2f} seconds"

#Starting our time for our full script runtime console printout
fullScritStart_Time = time.time()


#================================================================
#Writing our function for pulling the lat/long range for each project
import math

def get_lat_lng_bounds(lat, lng, radius_miles=3):
    """Calculate the bounding box for a given latitude, longitude, and radius in miles."""
    miles_per_degree_lat = 69.0  # Approximate miles per degree of latitude
    delta_lat = radius_miles / miles_per_degree_lat

    miles_per_degree_lng = 69.0 * math.cos(math.radians(lat))  # Adjust for latitude
    delta_lng = radius_miles / miles_per_degree_lng

    return {
        "min_lat": lat - delta_lat,
        "max_lat": lat + delta_lat,
        "min_lng": lng - delta_lng,
        "max_lng": lng + delta_lng,
    }


#================================================================
#Writing a function that will return the most frequently occuring item an a list: 
from collections import Counter

def most_frequent(lst):
    """
    Returns the most frequently occurring value in the list.
    If there are multiple values with the same highest frequency, returns one of them.
    """
    if not lst:
        return None  # Return None for empty lists
        
    counter = Counter(lst)
    return counter.most_common(1)[0][0]  # Get the most common value


#================================================================
#Writing a function that will pull the city out of our string that gets returned in our master gps data database table:
import re

def extract_city(address_string: str) -> str:
    """
    Extracts the city name from the given address string.
    The city is always the fourth element in the comma-separated address after 'GPS Coordinate Address:'.
    
    :param address_string: str, formatted address string
    :return: str, extracted city name
    """
    # Extract the part after 'GPS Coordinate Address:'
    match = re.search(r'GPS Coordinate Address: (.+)', address_string)
    
    if match:
        address_parts = [part.strip() for part in match.group(1).split(',')]
        
        # The city is the fourth element in the address structure
        if len(address_parts) >= 4:
            return address_parts[2]  # City name is at index 2 (zero-based index)
    
    return "City not found"  # Return a fallback message if extraction fails

# Example usage
#address_string = "OUTSIDE OF GEOFENCES! GPS Coordinate Address: 308, Plantation Drive, Coppell, Dallas County, Texas, 75019, United States"
#city = extract_city(address_string)


print('SUCCESS')

#endregion

#===============================================================================================================================================================
#Setting up our Supabase cloud database connection, logging in, AND creating some functions to use to access the data:
#region

print('CONNECTING TO OUR SUPABASE DATABASE, LOGGING IN, AND CREATING FUNCTIONS FOR PULLING/MODIFYING DATA...')

#=========================================================================
#Connecting to our Supabase cloud database:
#region

from supabase import create_client, Client

def connect_to_supabase(url: str, key: str) -> Client:
    """
    Connects to the Supabase database.

    Parameters:
    - url: The Supabase project URL.
    - key: The Supabase API key.

    Returns:
    - A Supabase client instance.
    """
    supabase: Client = create_client(url, key)
    return supabase

# Example usage:
supabase_url = 'https://tfaydxxaqmroiroazueg.supabase.co'
supabase_key = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InRmYXlkeHhhcW1yb2lyb2F6dWVnIiwicm9sZSI6ImFub24iLCJpYXQiOjE3Mzg3MDU4ODMsImV4cCI6MjA1NDI4MTg4M30.kBiVOV2loWuI_wnB3kmL0CE3jZOd6oOq02-bM3R8N-Y'
supabase_client = connect_to_supabase(supabase_url, supabase_key)

# Supabase credentials
supabase = create_client(supabase_url, supabase_key)


#endregion

#=========================================================================
#Logging in to our Supabase cloud database:
#region

#Function to log in a user
def login_user(email: str, password: str):
    response = supabase_client.auth.sign_in_with_password({"email": email, "password": password})
    return response

#Logging in to our Supabase databsae:
if __name__ == "__main__":
    # Log in as an authenticated user
    email = 'collingaines92@gmail.com'  # Replace with the user's email
    password = 'Cgained41$'   # Replace with the user's password
    login_response = login_user(email, password)

#endregion

#=========================================================================
#Writing some functions to use to access/edit our database later in this script:
#region

#================================
#Function for reading data from a specified database table. This will return a next list, similar to how you are used to dealing with Sqlite3:
def fetch_data_from_table(table_name: str):
    try:
        # Create a Supabase client
        supabase: Client = create_client(supabase_url, supabase_key)

        all_data = []
        limit = 1000  # Set the limit for each request
        offset = 0    # Start with an offset of 0

        while True:
            # Fetch data from the specified table with pagination
            response = supabase.table(table_name).select("*").range(offset, offset + limit - 1).execute()

            # Print the raw response for debugging
            #print("Raw response:", response)

            # Check if there was an error in the response
            # if response.error:
            #     print("Error fetching data:", response.error)
            #     return None

            # If no more data is returned, break the loop
            if not response.data:
                break

            # Append the fetched data to the all_data list
            all_data.extend(response.data)

            # Update the offset for the next request
            offset += limit

        # Convert the list of dictionaries to a nested list
        nested_list = [[value for value in row.values()] for row in all_data]

        return nested_list

    except Exception as e:
        print("An exception occurred:", str(e))
        return None

#================================
#Function for pulling data from our database WITH FILTERS
def fetch_filtered_data(supabase_url: str, supabase_key: str, table: str, filters: dict):
    """
    Fetch all data from a Supabase table where specified columns contain specified values.
    Returns data as a nested list (list of lists).

    :param supabase_url: Supabase project URL.
    :param supabase_key: Supabase API key.
    :param table: Name of the table to query.
    :param filters: Dictionary of column names and their required values.
    :return: Nested list of matching rows.
    """
    # Initialize Supabase client
    supabase: Client = create_client(supabase_url, supabase_key)
    
    # Start building the query
    query = supabase.table(table).select("*")

    # Apply filters
    for column, value in filters.items():
        query = query.eq(column, value)

    # Execute query
    response = query.execute()

    # Extract data properly
    if hasattr(response, 'data'):
        data = response.data  # Use .data attribute instead of .get()
    else:
        raise Exception(f"Unexpected Supabase response format: {response}")

    # Convert the list of dictionaries to a nested list
    nested_list = [list(row.values()) for row in data]

    return nested_list


# filters = {"column1": "value1", "column2": "value2"}
# results = fetch_filtered_data_nested(supabase_url, supabase_key, table_name, filters)
# print(results)

#================================
#Function for deleting rows from a database that have a specified value in a specified column: 
def delete_rows_by_value(supabase_url: str, supabase_key: str, table: str, column: str, value):
    """
    Deletes all rows from a specified Supabase table where a given column contains a specified value.

    Args:
        supabase_url (str): Your Supabase project URL.
        supabase_key (str): Your Supabase service role key.
        table (str): The name of the table from which to delete rows.
        column (str): The column to filter by.
        value: The value to match for deletion (type depends on the column).
    
    Returns:
        dict: The response from Supabase containing the status of the deletion.
    """
    try:
        # Create Supabase client
        supabase: Client = create_client(supabase_url, supabase_key)
        
        # Perform deletion
        response = supabase.table(table).delete().eq(column, value).execute()
        
        return response
    except Exception as e:
        return {"error": str(e)}

# Example Usage

# TABLE_NAME = "your_table_name"
# COLUMN_NAME = "your_column_name"
# VALUE_TO_DELETE = "your_value"

# result = delete_rows_with_value(supabase_url, supabase_key, "Master_Equipment_GPS_Data", "date", "2025-02-07")
# print(result)






#endregion


print('SUCCESS')
#endregion

#===============================================================================================================================================================
#Generating our access token for the HCSS API:
#region

print('CONNECTING TO OUR HCSS API...')

import requests

url = "https://api.hcssapps.com/identity/connect/token"

payload = {
    "client_id": "3i1ruzh3fhvfrrag38dkiqhnh4zp9wze",
    "client_secret": "NoK38HmOjmAj44blOXgsn2nxmPH0YCHiW02XIgbQ",
    "grant_type": "client_credentials",
    "scope": "e360:read e360:write heavyjob:read heavyjob:write timecards:read timecards:write heavyjob:users:read heavyjob:users:write dis:read dis:write skills:read skills:write telematics:read users:read users:write",
    "code": "string",
    "redirect_uri": "string"
}

headers = {"Content-Type": "application/x-www-form-urlencoded"}

response = requests.post(url, data=payload, headers=headers)

data = response.json()

token = data.get('access_token')

#Passing our token value generated above to our "HEADERS" variable:
HEADERS = {
"Authorization": "Bearer {}".format(token)
}

print('SUCCESS')

#endregion


#===============================================================================================================================================================
#Connecting to the Smartsheet API:
#region

print('CONNECTING TO OUR SMARTSHEET API...')

#Importing the Smartsheet library so that I can interact with it's API:
#SMARTSHEET API TOKEN (Collin's Application) ==> gFRPGyUEO4ykQlJQlmbrBqZiTmhbVCEuw8ol1
import smartsheet
import logging

#Initialize client. Uses the API token in the environment variable "SMARTSHEET_ACCESS_TOKEN"
smart = smartsheet.Smartsheet('gFRPGyUEO4ykQlJQlmbrBqZiTmhbVCEuw8ol1')

#Make sure we don't miss any errors:
smart.errors_as_exceptions(True)

print('SUCCESS')

#endregion


#endregion


#=====================================================================================================================================================================================================================================================================================================================
#=====================================================================================================================================================================================================================================================================================================================


#===============================================================================================================================================================
#First, let's create a list of all dates that we want to generate our report for. This list will always include all dates from Monday-Sunday of the current week, or if it is Monday it will generate a list of all dates for the prior week Mon-Sun:
#region

print('Creating a list of dates that we want to create our report for...')

from datetime import datetime, timedelta

def generate_week_dates():
    today = datetime.today()
    if today.weekday() == 0:  # Check if today is Monday
        monday_prior = today - timedelta(days=7)  # Get the Monday of the previous week
    else:
        monday_prior = today - timedelta(days=today.weekday())  # Get the most recent Monday
    
    week_dates = [(monday_prior + timedelta(days=i)).strftime('%Y-%m-%d') for i in range(7)]
    return week_dates


dates = generate_week_dates()

print('SUCCESS')
#endregion


#===============================================================================================================================================================
#Next, let's pull all of the the equipment utilization data from our "Master_Equipment_Utilization_Data" database and save it to a list of values to populate our excel report:
#region

#==================================================================================
#Creating a dictionary of all charge types for each pieced of equipment using our "Master Equipment List" smartsheet:

#============================================
#Adding equipment from our main asset list:
MySheet = smart.Sheets.get_sheet('1336754816634756')

chargeTypeDictionary = {}
statusDictionary = {}

for MyRow in MySheet.rows: 
    entryEquipID = MyRow.cells[0].value
    chargeType = MyRow.cells[11].value
    entryStatus = MyRow.cells[12].value

    chargeTypeDictionary[entryEquipID]=chargeType
    statusDictionary[entryEquipID]=entryStatus

#============================================
#Adding equipemtn from our fleet list:
MySheet = smart.Sheets.get_sheet('601782195539844')

for MyRow in MySheet.rows: 
    entryEquipID = MyRow.cells[0].value
    chargeType = MyRow.cells[11].value
    entryStatus = MyRow.cells[7].value

    chargeTypeDictionary[entryEquipID]=chargeType
    statusDictionary[entryEquipID]=entryStatus


#==================================================================================
#Compiling all of the data for each date into a single list:
equipmentUtilizationList = []


for i in range(len(dates)):
    entryDate = dates[i]

    #Pulling our data from our database:
    filters = {'date': entryDate}
    data = fetch_filtered_data(supabase_url, supabase_key, "Master_Equipment_Utilization_Data", filters)

    for j in range(len(data)):
        entryEquipID = data[j][2]
        entryEquipDescription = data[j][3]
        entryGPShours = data[j][4]
        heavyJobHours = data[j][5]
        hourDelta = data[j][6]
        primaryLocation = data[j][9]
        projectManager = data[j][12]
        foreman = data[j][11]


        #Updating our list:
        equipmentUtilizationList.append([entryDate, entryEquipID, entryEquipDescription, entryGPShours, heavyJobHours, hourDelta, primaryLocation, projectManager, foreman])



#endregion


#===============================================================================================================================================================
#Next, let's create our excel report that details the hour differences:
#region

#ITEMS TO ADD FOR UPDATE!
#> ONLY HIGHLIGHT RED HOUR DIFFERENCES GREATER THAN 0.25? MAYBE SEE HOW ACCURATE YOUR SYSTEM GPS DATA IS FIRST

#==================================================================================
#Creating our workbook/sheet objects:
wb = openpyxl.Workbook()
wb.save('EquipmentUtilizationReport.xlsx')

#Creating a sheet object: 
sheet = wb['Sheet']

#==================================================================================
#Clearing any previous values just in case our current report is shorter than a previous one:
row=1
for i in range(0,3000):
    sheet['A'+str(row)].value = ''
    sheet['B'+str(row)].value = ''
    sheet['C'+str(row)].value = ''
    sheet['D'+str(row)].value = ''
    sheet['E'+str(row)].value = ''
    sheet['F'+str(row)].value = ''
    sheet['G'+str(row)].value = ''
    sheet['H'+str(row)].value = ''
    sheet['I'+str(row)].value = ''

    row=row+1


#==================================================================================   
#Defining our report column headers and adding bold font as well as an underlined top row:

sheet['A1'].value = 'Date'
sheet['A1'].font = Font(bold=True)
sheet['B1'].value = 'Equipment ID'
sheet['B1'].font = Font(bold=True)
sheet['C1'].value = 'Equipment Description'
sheet['C1'].font = Font(bold=True)
sheet['D1'].value = 'GPS Hours'
sheet['D1'].font = Font(bold=True)
sheet['E1'].value = 'Heavy Job Hours'
sheet['E1'].font = Font(bold=True)
sheet['F1'].value = 'Delta'
sheet['F1'].font = Font(bold=True)
sheet['G1'].value = 'Primary Location'
sheet['G1'].font = Font(bold=True)
sheet['H1'].value = 'Project Manager'
sheet['H1'].font = Font(bold=True)
sheet['I1'].value = 'Foreman'
sheet['I1'].font = Font(bold=True)

#==================================================================================   
#Freezing the top row of our report and adding a double border to the bottom:

#Freezing the top row at cell C2 so that the date, equipID, and equip desc are frozen:
sheet.freeze_panes = "D2"

#Define a double border for the bottom
double_border = Border(bottom=Side(style='double'))

#Apply the border to each cell in the first row
for cell in sheet[1]:  # ws[1] selects the first row
    cell.border = double_border


#==================================================================================   
#Adding conditional formatting that highlights red any cells where the GPS hours don't match the HJ hours:

#Define the conditional formatting fill (red background)
red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

#Apply conditional formatting to a specified column (e.g., Column C)
column_to_format = "F"  # Change this to your desired column

sheet.conditional_formatting.add(
    f"{column_to_format}2:{column_to_format}10000",  # Applies from row 2 to 100
    CellIsRule(operator="greaterThan", formula=["0.25"], fill=red_fill)
)

sheet.conditional_formatting.add(
    f"{column_to_format}2:{column_to_format}10000",
    CellIsRule(operator="lessThan", formula=["-0.25"], fill=red_fill)
)


#==================================================================================
#Populating the cells in our excel spreadsheet: 
row=2

#equipmentUtilizationList.append([entryDate, entryEquipID, entryEquipDescription, round(entryGPShours,2), round(heavyJobHours,2), round(hourDelta,2), primaryLocation, 'Project Manager', foreman])
for i in range(len(equipmentUtilizationList)):
    sheet['A'+str(row)].value = equipmentUtilizationList[i][0]
    sheet['B'+str(row)].value = equipmentUtilizationList[i][1]
    sheet['C'+str(row)].value = equipmentUtilizationList[i][2]
    sheet['D'+str(row)].value = equipmentUtilizationList[i][3]
    sheet['E'+str(row)].value = equipmentUtilizationList[i][4]
    sheet['F'+str(row)].value = equipmentUtilizationList[i][5]
    sheet['G'+str(row)].value = equipmentUtilizationList[i][6]
    sheet['H'+str(row)].value = equipmentUtilizationList[i][7]
    sheet['I'+str(row)].value = equipmentUtilizationList[i][8]

    row=row+1



#==================================================================================
#Finally, let's save the workbook:

print(equipmentUtilizationList)

wb.save('EquipmentUtilizationReport.xlsx')


#endregion


#===============================================================================================================================================================
#Finally, sending our email:
#region

print('Generating email and sending to relevant DDM employees...')

#==================================================================================
#First, let's use our "dates" list defined at the top of this script to return the first/last date that this report is covering:
startDate = dates[0]
endDate = dates[-1]


#==================================================================================
#Next, building the email:
import smtplib
from email.message import EmailMessage

conn = smtplib.SMTP('smtp-mail.outlook.com', 587)
#the ehlo function actually connects you to the server
conn.ehlo()
#starting the tls encryption to protect our password
conn.starttls()
#now that we are connected, let's login
conn.login('automatedreporting@ddmcc.net', 'CG@ddm92')

newMessage = EmailMessage()    #creating an object of EmailMessage class
newMessage['Subject'] = "Equipment Hour Correction Audit for {} through {}".format(startDate, endDate) #Defining email subject
newMessage['From'] = 'automatedreporting@ddmcc.net'  #Defining sender email
#newMessage['To'] = 'jroden@ddmcc.net, rbeltran@ddmcc.net, tyoes@ddmcc.net, gtrabazo@ddmcc.net, collin@ddmcc.net, bkuecker@ddmcc.net, zack@ddmcc.net, mruez@ddmcc.net, croberts@ddmcc.net, jderiso@ddmcc.net, bpoeschl@ddmcc.net, rlow@ddmcc.net, msoto@ddmcc.net, ggonzalez@ddmcc.net, fcoutee@ddmcc.net, RBandeira@ddmcc.net, rreyes@ddmcc.net, MBartos@ddmcc.net, Abernard@ddmcc.net, DDodson@ddmcc.net, droot@ddmcc.net'  #Defining reciever email
newMessage['To'] = 'collin@ddmcc.net, pyramidconstructionsupply@outlook.com'  #Defining reciever email

#Now let's build the string to be added to our email message!
emailcontentstring = 'Hello DDM Team,\n\nPlease see attached for the updated equipment hour audit report for {} through {}. Please make any and all listed corrections as soon as possible.\n\nThis is an automated email, so any direct replies may go unread. If you have any questions, please hit "Reply All" or make sure that Derek, Blake, and Collin are included on your email response.'.format(startDate, endDate)

newMessage.set_content(emailcontentstring)


#==================================================================================
#Finally, attaching our excel report and sending this email

#Converting the excel file into binary format so that we can attach it to our email:
def convert_into_binary(file_path):
    with open(file_path, 'rb') as file:
        binary = file.read()
    return binary

excelreport = convert_into_binary("EquipmentUtilizationReport.xlsx")

#Adding the excel report as an attachment:
newMessage.add_attachment(excelreport, maintype='application', subtype='pdf', filename='Equipment Hour Correction Audit ({} through {}).xlsx'.format(startDate, endDate))

#Sending the email:
conn.send_message(newMessage)

#When you're done just call the quit method to end the connection
conn.quit()


print('SUCCESS')


#endregion